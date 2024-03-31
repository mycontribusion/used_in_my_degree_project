import openpyxl

# Specify the Excel file path to be edited
excel_file_path = 'Responses.xlsx'

# Load the existing Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Select the sheet you want to edit (assuming the first sheet for simplicity)
sheet = workbook.active

one=["10-13","Female","Single","Islam","Hausa","Option 1","Confident","Oct-13","10/1/2013","None","House wife","Yes","Health facility","Potential side effect","Prevention of cervical cancer","Very confident","Friend","Very acceptable","Free vaccination","Fear of side effect","Fear of side effects","Fear of side effects, Predispose to sexual activity"]
two=["14-17","Married","Christianity","Fulani","Quranic","Option 2","Trader","No","Family","Effectiveness of the vaccine","Reduced risk of genital warts","Not confident","Health worker","Acceptable","Going home by home","Cost"]
three=["18-19","Igbo","Primary","Civil servant","Media","Cost of vaccine","Not sure","Family member","Not acceptable","Giving free gift","Predispose to sexual activity"]
four=["Yoruba","secondary","Student","student ","Student ","I don't care","Nothing","I don't care"]
five=["Tertiary"]

# Iterate through all rows and columns
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        # Example: Change values based on some condition
        if cell.value is not None and cell.value in one:
            cell.value = 1

        if cell.value is not None and cell.value in two:
            cell.value = 2

        if cell.value is not None and cell.value in three:
            cell.value = 3

        if cell.value is not None and cell.value in four:
            cell.value = 4

        if cell.value is not None and cell.value in five:
            cell.value = 5

# Save the changes to the Excel file
workbook.save('Responses.xlsx')

print('Values edited and saved successfully.')
