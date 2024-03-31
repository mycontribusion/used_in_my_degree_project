# import openpyxl module
import openpyxl
 
# Give the location of the file
path = "project edited py.xlsx"
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
cell_obj = sheet_obj.cell(row=1, column=1)
 
print(cell_obj.value)
 
row = sheet_obj.max_row
column = sheet_obj.max_column
 
print("Total Rows:", row)
print("Total Columns:", column)
 
#print("\nValue of first column")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)
 
#print("\nValue of first row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")
