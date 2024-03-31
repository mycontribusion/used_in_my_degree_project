import openpyxl

# Specify the Excel file path to be edited
excel_file_path = 'project with L2.xlsx'

# Load the existing Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Select the sheet you want to edit (assuming the first sheet for simplicity)
sheet = workbook.active

one=["Male", "Under 20", "Islam", "Hausa", "Never","Not at all", "Not satisfied at all","Yes","Too low"]
two=["Female","20-24","Christianity","Igbo","Rarely","Slightly","No","Maybe","Slight low","Slightly satisfied"]
three=["25-30","Yoruba","Occasionally","Moderately","Indifferent","Okay","Moderately satisfied"]
four=["Above 30","Frequently","Very much","Slightly excessive","Very satisfied","Often"]
five=["Always","Extremely","Too excessive","Extremely satisfied"]

# Iterate through all rows and columns
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        # Example: Change values based on some condition
        if cell.value is not None and cell.value in one:
            cell.value = 1

        if cell.value is not None and cell.value in two:
            cell.value = 2

        if cell.value is not None and cell.value in three:
            cell.value = 3

        if cell.value is not None and cell.value in four:
            cell.value = 4

        if cell.value is not None and cell.value in five:
            cell.value = 5

# Save the changes to the Excel file
workbook.save('project with L2.xlsx')

print('Values edited and saved successfully.')
