import openpyxl

# Specify the Excel file path to be edited
excel_file_path = 'book9.xlsx'

# Load the existing Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Select the sheet you want to edit (assuming the first sheet for simplicity)
sheet = workbook.active

one=["Male", "Under 20", "Islam", "Hausa", "Never","Not at all", "Not satisfied at all","Yes","Too low"]
two=["Female","20-24","Christianity","Igbo","Rarely","Slightly","No","Maybe","Slight low","Slightly satisfied"]
three=["25-30","Yoruba","Occasionally","Moderately","Indifferent","Okay","Moderately satisfied"]
four=["Above 30","Frequently","Very much","Slightly excessive","Very satisfied","Often"]
five=["Always","Extremely","Too excessive","Extremely satisfied"]

#low=["0","1","2","3","4","5","6","7","8"]
#moderate=["9","10","11","12","13","14","15","16"]
#high=["17","18","19","20","21","22","23","24"]

low=[0,1,2,3,4,5,6,"Low"]
moderate=[7,8,9,10,11,12,13,14,15,"Moderate"]
high=[16,17,18,19,20,21,22,23,24,"High"]



# Iterate through all rows and columns
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

    for cell in row:        
        # Example: Change values based on some condition
        if cell.value <20:
            cell.value = 1

        if cell.value >=20 and cell.value <25:
            cell.value = 2

        if cell.value <=29 and cell.value >=25:
            cell.value = 3

        if cell.value > 29:
            cell.value = 4

        #if cell.value is not None and cell.value in one:
            #cell.value = 4

        #if cell.value is not None and cell.value in two:
            #cell.value = 3

        #if cell.value is not None and cell.value in three:
            #cell.value = 2

        #if cell.value is not None and cell.value in four:
            #cell.value = 1

        #if cell.value is not None and cell.value in five:
            #cell.value = 0
            

# Save the changes to the Excel file
workbook.save('book9.xlsx')

print('Values edited and saved successfully.')
